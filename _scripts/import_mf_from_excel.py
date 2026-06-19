"""
_scripts/import_mf_from_excel.py
==================================
Parses MF portfolio Excel files (any AMC format) and saves to Neon.
Works with 360 ONE, HDFC, SBI, Axis, any AMC that uses AMFI standard format.

AMFI standard format (all AMCs use this):
  Col A: Internal code
  Col B: Name of instrument  
  Col C: ISIN
  Col D: Industry/Rating
  Col E: Quantity
  Col F: Market/Fair Value (Rs. in Lacs)
  Col G: % to Net Assets

Usage:
  python _scripts/import_mf_from_excel.py --dir /path/to/excel/files
  python _scripts/import_mf_from_excel.py --file portfolio_may2026.xlsx
  python _scripts/import_mf_from_excel.py --dir C:/Downloads/amfi_portfolios --amc "SBI MF"
"""

import os, sys, re, logging, argparse, datetime
import openpyxl, psycopg2, psycopg2.extras
from collections import defaultdict

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
log = logging.getLogger(__name__)

DATABASE_URL = os.environ.get("DATABASE_URL") or os.environ.get("NEON_DATABASE_URL")

# ── Name → NSE symbol mapping ─────────────────────────────────────────────────
# Extend this as more stocks are encountered
NAME_MAP = {
    "HDFC Bank":               "HDFCBANK",
    "ICICI Bank":              "ICICIBANK",
    "Infosys":                 "INFY",
    "Tata Motors":             "TATAMOTORS",
    "NTPC":                    "NTPC",
    "Indus Towers":            "INDUSTOWER",
    "Cholamandalam":           "CHOLAFIN",
    "Bharti Airtel":           "BHARTIARTL",
    "Larsen & Toubro":         "LT",
    "Motherson Sumi":          "MOTHERSON",
    "Coal India":              "COALINDIA",
    "Axis Bank":               "AXISBANK",
    "APL Apollo":              "APLAPOLLO",
    "Cummins India":           "CUMMINSIND",
    "Divi's":                  "DIVISLAB",
    "Dixon Technologies":      "DIXON",
    "Tech Mahindra":           "TECHM",
    "REC Limited":             "RECLTD",
    "Hindustan Aeronautics":   "HAL",
    "Oracle Financial":        "OFSS",
    "Hero MotoCorp":           "HEROMOTOCO",
    "Macrotech":               "LODHA",
    "Bharat Electronics":      "BEL",
    "LIC Housing":             "LICHSGFIN",
    "TVS Motor":               "TVSMOTOR",
    "Zydus":                   "ZYDUSLIFE",
    "Solar Industries":        "SOLARINDS",
    "Power Finance":           "PFC",
    "Power Grid":              "POWERGRID",
    "Bajaj Auto":              "BAJAJ-AUTO",
    "InterGlobe":              "INDIGO",
    "Torrent Pharma":          "TORNTPHARM",
    "Aurobindo":               "AUROPHARMA",
    "ICICI Lombard":           "ICICIGI",
    "Reliance Industries":     "RELIANCE",
    "Tata Consultancy":        "TCS",
    "Wipro":                   "WIPRO",
    "HCL Tech":                "HCLTECH",
    "State Bank":              "SBIN",
    "Kotak Mah":               "KOTAKBANK",
    "Sun Pharma":              "SUNPHARMA",
    "Titan":                   "TITAN",
    "Bajaj Finance":           "BAJFINANCE",
    "Maruti":                  "MARUTI",
    "Dr Reddy":                "DRREDDY",
    "Cipla":                   "CIPLA",
    "Sona BLW":                "SONACOMS",
    "Samvardhana Motherson":   "MOTHERSON",
    "Aavas Financiers":        "AAVAS",
    "Sumitomo Chemical":       "SUMICHEM",
    "Blue Dart":               "BLUEDART",
    "Netweb":                  "NETWEB",
    "RBL Bank":                "RBLBANK",
    "Federal Bank":            "FEDERALBNK",
    "AU Small Finance":        "AUBANK",
    "IDFC First":              "IDFCFIRSTB",
    "Bandhan":                 "BANDHANBNK",
    "Tata Steel":              "TATASTEEL",
    "JSW Steel":               "JSWSTEEL",
    "Hindalco":                "HINDALCO",
    "Grasim":                  "GRASIM",
    "UltraTech":               "ULTRACEMCO",
    "Asian Paints":            "ASIANPAINT",
    "Berger Paints":           "BERGEPAINT",
    "Pidilite":                "PIDILITIND",
    "Havells":                 "HAVELLS",
    "Crompton":                "CROMPTON",
    "Voltas":                  "VOLTAS",
    "Polycab":                 "POLYCAB",
    "KEI Industries":          "KEI",
    "Varun Beverages":         "VBL",
    "United Spirits":          "UNITDSPR",
    "Britannia":               "BRITANNIA",
    "Nestle":                  "NESTLEIND",
    "Hindustan Unilever":      "HINDUNILVR",
    "ITC":                     "ITC",
    "Marico":                  "MARICO",
    "Dabur":                   "DABUR",
    "Godrej Consumer":         "GODREJCP",
    "Colgate":                 "COLPAL",
    "Avenue Supermarts":       "DMART",
    "Trent":                   "TRENT",
    "Nykaa":                   "NYKAA",
    "Zomato":                  "ZOMATO",
    "Paytm":                   "PAYTM",
    "SBI Life":                "SBILIFE",
    "HDFC Life":               "HDFCLIFE",
    "Max Financial":           "MFSL",
    "Bajaj Finserv":           "BAJAJFINSV",
    "Muthoot Finance":         "MUTHOOTFIN",
    "Shriram Finance":         "SHRIRAMFIN",
    "Piramal":                 "PIRAMAL",
    "Aditya Birla Capital":    "ABCAPITAL",
    "Tata Power":              "TATAPOWER",
    "Adani Green":             "ADANIGREEN",
    "Adani Ports":             "ADANIPORTS",
    "JSW Energy":              "JSWENERGY",
    "Torrent Power":           "TORNTPOWER",
    "CESC":                    "CESC",
    "Apollo Hospitals":        "APOLLOHOSP",
    "Max Healthcare":          "MAXHEALTH",
    "Fortis":                  "FORTIS",
    "Narayana":                "NH",
    "Dr Lal PathLabs":         "LALPATHLAB",
    "Metropolis":              "METROPOLIS",
    "Persistent":              "PERSISTENT",
    "Mphasis":                 "MPHASIS",
    "Coforge":                 "COFORGE",
    "L&T Technology":          "LTTS",
    "Mindtree":                "MINDTREE",
    "Cyient":                  "CYIENT",
    "Tata Elxsi":              "TATAELXSI",
    "Astral":                  "ASTRAL",
    "Finolex":                 "FINPIPE",
    "Supreme Industries":      "SUPREMEIND",
    "PI Industries":           "PIIND",
    "UPL":                     "UPL",
    "Bayer CropScience":       "BAYERCROP",
    "Rallis":                  "RALLIS",
    "Abbott India":            "ABBOTINDIA",
    "Alkem":                   "ALKEM",
    "Lupin":                   "LUPIN",
    "Biocon":                  "BIOCON",
    "Gland Pharma":            "GLAND",
    "Granules":                "GRANULES",
    "Neuland":                 "NEULANDLAB",
    "SRF":                     "SRF",
    "Navin Fluorine":          "NAVINFLUOR",
    "Deepak Nitrite":          "DEEPAKNTR",
    "Aarti Industries":        "AARTIIND",
    "Fine Organic":            "FINEORG",
    "Linde India":             "LINDEINDIA",
    "Schaeffler":              "SCHAEFFLER",
    "Timken":                  "TIMKEN",
    "SKF":                     "SKFINDIA",
    "Bharat Forge":            "BHARATFORG",
    "Ramkrishna Forgings":     "RKFORGE",
    "Mahindra & Mahindra":     "M&M",
    "Eicher Motors":           "EICHERMOT",
    "Bosch":                   "BOSCHLTD",
    "Minda Industries":        "UNOMINDA",
    "Endurance Tech":          "ENDURANCE",
    "Sansera Engineering":     "SANSERA",
    "Amber Enterprises":       "AMBER",
    "Voltas":                  "VOLTAS",
    "Blue Star":               "BLUESTARCO",
    "Whirlpool":               "WHIRLPOOL",
    "Bata India":              "BATAINDIA",
    "Page Industries":         "PAGEIND",
    "Vedant Fashions":         "MANYAVAR",
    "Go Fashion":              "GOCOLORS",
    "PVR":                     "PVRINOX",
    "Inox Leisure":            "PVRINOX",
    "Thomas Cook":             "THOMASCOOK",
    "Indian Hotels":           "INDHOTEL",
    "EIH":                     "EIHOTEL",
    "Lemon Tree":              "LEMONTREE",
    "DLF":                     "DLF",
    "Godrej Properties":       "GODREJPROP",
    "Prestige Estates":        "PRESTIGE",
    "Oberoi Realty":           "OBEROIRLTY",
    "Brigade":                 "BRIGADE",
    "Sobha":                   "SOBHA",
    "Phoenix Mills":           "PHOENIXLTD",
    "Embassy":                 "EMBASSY",
    "Mindspace":               "MINDSPACE",
}

EQUITY_SHEET_KEYWORDS = ['equity', 'flexicap', 'quant', 'hybrid', 'elss', 'growth',
                          'large', 'mid', 'small', 'multi', 'value', 'focused', 'flexi']
SKIP_NAMES = {'Sub Total', 'Total', 'Grand Total', 'NIL', '', 'Name of the Instrument'}
SKIP_KEYWORDS = ['government of india', 'treasury bill', 'certificate of deposit',
                 'commercial paper', 'reit', 'invit', 'money market', 'sovereign',
                 'state government', '% government', 'short term deposit']

def find_symbol(name: str) -> str | None:
    nl = name.lower()
    for key, sym in NAME_MAP.items():
        if key.lower() in nl:
            return sym
    return None

def extract_month(filename: str) -> str:
    m = re.search(r'(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[\s_-]?(\d{4})', 
                  filename, re.I)
    if m:
        month = m.group(1).capitalize()[:3]
        year  = m.group(2)
        return f"{month}{year}"
    return datetime.date.today().strftime("%b%Y")

def extract_amc(filename: str, default: str) -> str:
    fname = os.path.basename(filename).upper()
    if '360' in fname:    return "360 ONE"
    if 'HDFC' in fname:   return "HDFC MF"
    if 'SBI' in fname:    return "SBI MF"
    if 'AXIS' in fname:   return "Axis MF"
    if 'KOTAK' in fname:  return "Kotak MF"
    if 'NIPPON' in fname: return "Nippon MF"
    if 'MIRAE' in fname:  return "Mirae Asset"
    if 'ICICI' in fname:  return "ICICI Pru MF"
    if 'UTI' in fname:    return "UTI MF"
    if 'DSP' in fname:    return "DSP MF"
    if 'MOTILAL' in fname:return "Motilal Oswal MF"
    return default

def parse_file(filepath: str, amc_override: str = "") -> dict:
    """Parse one AMC portfolio Excel file. Returns {symbol: [{scheme,amc,value,weight}]}"""
    holdings: dict = defaultdict(list)
    month_str = extract_month(filepath)
    amc_name  = amc_override or extract_amc(filepath, "Unknown AMC")
    
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    except Exception as e:
        log.warning(f"Cannot open {os.path.basename(filepath)}: {e}")
        return {}
    
    for sheet_name in wb.sheetnames:
        if not any(k in sheet_name.lower() for k in EQUITY_SHEET_KEYWORDS):
            continue
        
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        
        # Get scheme name from row 0
        scheme_name = ""
        if rows and rows[0]:
            for cell in rows[0]:
                if cell and len(str(cell)) > 5:
                    scheme_name = str(cell)[:100]
                    break
        
        for row in rows[3:]:
            if not row or len(row) < 6:
                continue
            
            name      = str(row[1] or '').strip()
            value_raw = row[5]
            weight_raw= row[6] if len(row) > 6 else None
            
            if not name or name in SKIP_NAMES:
                continue
            if any(kw in name.lower() for kw in SKIP_KEYWORDS):
                continue
            
            try:
                value_cr  = float(value_raw or 0) / 100
                weight_pct= float(weight_raw or 0) * 100
            except:
                continue
            
            if value_cr <= 0:
                continue
            
            sym = find_symbol(name)
            if sym:
                holdings[sym].append({
                    "month":   month_str,
                    "scheme":  scheme_name or sheet_name,
                    "amc":     amc_name,
                    "value":   round(value_cr, 2),
                    "weight":  round(weight_pct, 4),
                })
    
    return dict(holdings)

def get_db():
    return psycopg2.connect(DATABASE_URL, connect_timeout=15)

def ensure_tables(conn):
    cur = conn.cursor()
    cur.execute("""
        CREATE TABLE IF NOT EXISTS mf_stock_summary (
            id SERIAL PRIMARY KEY, nse_symbol TEXT NOT NULL,
            month DATE NOT NULL, fund_count INTEGER, amc_count INTEGER,
            net_additions INTEGER DEFAULT 0, net_exits INTEGER DEFAULT 0,
            accumulation_score NUMERIC(6,2),
            UNIQUE(nse_symbol, month)
        );
        ALTER TABLE mf_stock_summary ADD COLUMN IF NOT EXISTS total_value_cr NUMERIC(14,2);
        ALTER TABLE mf_stock_summary ADD COLUMN IF NOT EXISTS signal TEXT;
        CREATE TABLE IF NOT EXISTS mf_scheme_holdings (
            id SERIAL PRIMARY KEY, nse_symbol TEXT NOT NULL,
            month DATE NOT NULL, amc_name TEXT, scheme_name TEXT,
            market_value_cr NUMERIC(14,2), portfolio_weight_pct NUMERIC(6,2),
            UNIQUE(nse_symbol, month, scheme_name)
        );
    """)
    conn.commit()
    cur.close()

def month_to_date(month_str: str) -> datetime.date:
    try:
        return datetime.datetime.strptime(month_str, "%b%Y").date().replace(day=1)
    except:
        return datetime.date.today().replace(day=1)

def save_holdings(conn, all_holdings: dict):
    """Save aggregated holdings across all files to Neon."""
    # all_holdings: {symbol: [{month, scheme, amc, value, weight}]}
    cur = conn.cursor()
    saved_stocks = 0; saved_schemes = 0

    # Group by symbol + month
    by_sym_month: dict = defaultdict(lambda: defaultdict(list))
    for sym, records in all_holdings.items():
        for r in records:
            by_sym_month[sym][r['month']].append(r)

    for sym, months in by_sym_month.items():
        for month_str, records in months.items():
            month_date  = month_to_date(month_str)
            total_value = sum(r['value'] for r in records)
            fund_count  = len(records)
            amc_count   = len(set(r['amc'] for r in records))
            signal      = ("HEAVY_BUYING" if total_value > 5000 else
                           "MODERATE"     if total_value > 500  else "LIGHT")

            cur.execute("""
                INSERT INTO mf_stock_summary
                    (nse_symbol, month, total_value_cr, fund_count, amc_count, signal)
                VALUES (%s,%s,%s,%s,%s,%s)
                ON CONFLICT (nse_symbol, month) DO UPDATE SET
                    total_value_cr = COALESCE(mf_stock_summary.total_value_cr,0) + EXCLUDED.total_value_cr,
                    fund_count     = mf_stock_summary.fund_count + EXCLUDED.fund_count,
                    amc_count      = mf_stock_summary.amc_count + EXCLUDED.amc_count,
                    signal         = EXCLUDED.signal
            """, (sym, month_date, round(total_value,2), fund_count, amc_count, signal))
            saved_stocks += 1

            # Deduplicate by scheme name before inserting
            seen_schemes = set()
            deduped = []
            for r in records:
                key = r['scheme'][:200]
                if key not in seen_schemes:
                    seen_schemes.add(key)
                    deduped.append(r)

            cur.execute("DELETE FROM mf_scheme_holdings WHERE nse_symbol=%s AND month=%s",
                        (sym, month_date))
            for r in deduped[:20]:
                cur.execute("""
                    INSERT INTO mf_scheme_holdings
                        (nse_symbol, month, amc_name, scheme_name, market_value_cr, portfolio_weight_pct)
                    VALUES (%s,%s,%s,%s,%s,%s)
                """, (sym, month_date, r['amc'][:100], r['scheme'][:200],
                      r['value'], r['weight']))
                saved_schemes += 1

    conn.commit()
    cur.close()
    return saved_stocks, saved_schemes

def main():
    p = argparse.ArgumentParser()
    p.add_argument("--dir",  help="Directory with Excel files")
    p.add_argument("--file", help="Single Excel file")
    p.add_argument("--amc",  default="", help="AMC name override")
    args = p.parse_args()

    if not DATABASE_URL:
        log.error("DATABASE_URL not set"); sys.exit(1)

    conn = get_db()
    log.info("Connected to DB")
    ensure_tables(conn)

    # Collect files
    files = []
    if args.file:
        files = [args.file]
    elif args.dir:
        for f in os.listdir(args.dir):
            if f.endswith(('.xlsx','.xls','.XLS')):
                files.append(os.path.join(args.dir, f))
    else:
        log.error("Provide --dir or --file"); sys.exit(1)

    log.info(f"Processing {len(files)} files")
    log.info("=" * 60)

    all_holdings: dict = defaultdict(list)

    for fpath in sorted(files):
        fname = os.path.basename(fpath)
        if 'Overlap' in fname or '2018' in fname:
            continue
        
        holdings = parse_file(fpath, args.amc)
        for sym, records in holdings.items():
            all_holdings[sym].extend(records)
        
        total = sum(len(v) for v in holdings.values())
        if total > 0:
            log.info(f"  ✓ {fname}: {len(holdings)} stocks, {total} holdings")

    stocks, schemes = save_holdings(conn, all_holdings)
    conn.close()

    log.info("=" * 60)
    log.info(f"Done. {stocks} stock-months, {schemes} scheme rows saved to Neon")
    log.info("MF Intelligence panel will now show data for these stocks.")

if __name__ == "__main__":
    main()
