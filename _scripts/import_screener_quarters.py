#!/usr/bin/env python3
"""
import_screener_quarters.py — capture the **Quarters** section of the Screener Data Sheet (full
quarterly P&L: Sales, Expenses, Operating Profit, OPM%, Other Income, Depreciation, Interest, PBT,
Tax%, Net Profit, EPS) into a new `quarterly_financials` table. Reuses the *_10yr.xlsx files the
auto-refresh already downloads — NO new scraping. Separate from import_screener_financials.py so
the working annual import is never at risk.

Why: the annual importer parses *past* the Quarters / PRICE: / DERIVED: sections and discards them.
Quarters unlocks the earnings-surprise engine (full quarterly P&L vs revenue/PAT only) and quarterly
trend charts.

Modes:
  --inspect FILE     dump every row label under Quarters / PRICE: / DERIVED: of ONE file, then exit.
                     (run this once on a real file and paste output so PRICE:/DERIVED: get exact maps)
  --diag             parse one file's Quarters, print it, flag unmapped labels, NO DB writes
  (default)          parse --dir and upsert quarterly_financials

  python _scripts/import_screener_quarters.py --inspect data/fundamental_raw/RELIANCE_10yr.xlsx
  python _scripts/import_screener_quarters.py --dir data/fundamental_raw --diag
  python _scripts/import_screener_quarters.py --dir data/fundamental_raw
"""
import os, sys, glob, argparse, datetime
import openpyxl

URL = os.environ.get("DATABASE_URL") or os.environ.get("NEON_DATABASE_URL")

# Quarters section labels (col A) -> db column. Confirmed against the live Data Sheet:
# Tax is a rupee AMOUNT (not Tax %), and there is no OPM% or EPS row (OPM% is computed below).
QMAP = {
    "Sales": "sales", "Sales+": "sales",
    "Expenses": "expenses", "Expenses+": "expenses",
    "Operating Profit": "operating_profit",
    "Other Income": "other_income", "Other Income+": "other_income",
    "Depreciation": "depreciation",
    "Interest": "interest",
    "Profit before tax": "pbt",
    "Tax": "tax",
    "Net Profit": "net_profit", "Net profit": "net_profit", "Net Profit+": "net_profit",
}
QCOLS = ["sales", "expenses", "operating_profit", "opm_pct", "other_income",
         "depreciation", "interest", "pbt", "tax", "net_profit"]
HEADERS = ["META", "PROFIT & LOSS", "Quarters", "BALANCE SHEET", "CASH FLOW:", "PRICE:", "DERIVED:"]


def num(v):
    return float(v) if isinstance(v, (int, float)) else None


def _load_rows(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    if "Data Sheet" not in wb.sheetnames:
        return None
    ws = wb["Data Sheet"]
    return [[ws.cell(r, c).value for c in range(1, ws.max_column + 1)] for r in range(1, ws.max_row + 1)]


def _bounds(rows, header):
    for i, rr in enumerate(rows):
        if rr and rr[0] == header:
            return i
    return None


def _section_range(rows, header):
    """[start, end) for a section — end = the next recognized header below it."""
    start = _bounds(rows, header)
    if start is None:
        return None, None
    positions = sorted(p for h in HEADERS for p in [_bounds(rows, h)] if p is not None and p > start)
    end = positions[0] if positions else len(rows)
    return start, end


def parse_quarters(path):
    """Return (symbol_name, [ {period: date, fiscal_label, <cols>} ... ]) from the Quarters section."""
    rows = _load_rows(path)
    if rows is None:
        return None, [], []
    company = next((rr[1] for rr in rows if rr and rr[0] == "COMPANY NAME"), None)
    start, end = _section_range(rows, "Quarters")
    if start is None:
        return company, [], []

    # period columns from the Report Date row inside the Quarters block
    dates = None
    for i in range(start, end):
        if rows[i] and rows[i][0] == "Report Date":
            dates = rows[i][1:]
            break
    if not dates:
        return company, [], []

    periods = [d if isinstance(d, datetime.datetime) else None for d in dates]
    cols = {j: {} for j in range(len(periods)) if periods[j] is not None}
    unmapped = []
    for i in range(start, end):
        label = rows[i][0] if rows[i] else None
        if label in (None, "Report Date", "Quarters"):
            continue
        col = QMAP.get(label)
        if col is None:
            if isinstance(label, str) and label.strip():
                unmapped.append(label)
            continue
        vals = rows[i][1:]
        for j in cols:
            if j < len(vals):
                cols[j][col] = num(vals[j])

    out = []
    for j, rec in cols.items():
        d = periods[j]
        rec["period"] = d.date()
        rec["fiscal_label"] = f"{d.year}-{d.month:02d}"
        # OPM% isn't a row in the Data Sheet's Quarters block — derive it.
        s, op = rec.get("sales"), rec.get("operating_profit")
        rec["opm_pct"] = round(op / s * 100, 2) if (s and op is not None and s != 0) else None
        out.append(rec)
    out.sort(key=lambda r: r["period"])
    return company, out, sorted(set(unmapped))


def inspect(path):
    rows = _load_rows(path)
    if rows is None:
        sys.exit("No 'Data Sheet' tab in that file.")
    for header in ("Quarters", "PRICE:", "DERIVED:"):
        s, e = _section_range(rows, header)
        print(f"\n===== {header} =====")
        if s is None:
            print("  (section not present)")
            continue
        for i in range(s, e):
            label = rows[i][0] if rows[i] else None
            if label is None:
                continue
            sample = [x for x in (rows[i][1:4] if rows[i] else []) if x is not None][:3]
            print(f"  {str(label):28} e.g. {sample}")


def ensure_table(cur):
    cur.execute("""
        CREATE TABLE IF NOT EXISTS quarterly_financials (
            symbol TEXT NOT NULL,
            period DATE NOT NULL,
            PRIMARY KEY (symbol, period))""")
    # Evolve the schema so a table created by an EARLIER version of this script (which had
    # tax_pct/eps but no `tax`) gains the current columns. ADD COLUMN IF NOT EXISTS is a no-op
    # where the column already exists, so this is safe on fresh and pre-existing tables alike.
    coltypes = {
        "fiscal_label": "TEXT",
        "sales": "NUMERIC(18,4)", "expenses": "NUMERIC(18,4)", "operating_profit": "NUMERIC(18,4)",
        "opm_pct": "NUMERIC(10,4)", "other_income": "NUMERIC(18,4)", "depreciation": "NUMERIC(18,4)",
        "interest": "NUMERIC(18,4)", "pbt": "NUMERIC(18,4)", "tax": "NUMERIC(18,4)",
        "net_profit": "NUMERIC(18,4)", "updated_at": "TIMESTAMPTZ DEFAULT NOW()",
    }
    for c, t in coltypes.items():
        cur.execute(f"ALTER TABLE quarterly_financials ADD COLUMN IF NOT EXISTS {c} {t}")


def symbol_from_path(path):
    return os.path.basename(path).replace("_10yr.xlsx", "").replace(".xlsx", "").upper()


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--dir")
    ap.add_argument("--inspect", dest="inspect_file")
    ap.add_argument("--diag", action="store_true")
    args = ap.parse_args()

    if args.inspect_file:
        inspect(args.inspect_file); return
    if not args.dir:
        sys.exit("Pass --dir DIR (or --inspect FILE)")

    files = sorted(glob.glob(os.path.join(args.dir, "*.xlsx")))
    if not files:
        sys.exit(f"No .xlsx in {args.dir}")

    if args.diag:
        sym = symbol_from_path(files[0])
        company, quarters, unmapped = parse_quarters(files[0])
        print(f"{sym} ({company}) — {len(quarters)} quarters")
        for q in quarters[-8:]:
            print(f"  {q['fiscal_label']}: sales={q.get('sales')}, op={q.get('operating_profit')}, "
                  f"opm%={q.get('opm_pct')}, net={q.get('net_profit')}, tax={q.get('tax')}")
        print(f"  unmapped labels in Quarters: {unmapped or 'none'}")
        return

    import psycopg2
    from psycopg2.extras import execute_values
    if not URL:
        sys.exit("DATABASE_URL not set")
    conn = psycopg2.connect(URL); cur = conn.cursor()
    ensure_table(cur); conn.commit()

    rows_out, n_q, n_files = [], 0, 0
    for path in files:
        sym = symbol_from_path(path)
        try:
            _, quarters, _ = parse_quarters(path)
        except Exception as e:
            print(f"  skip {sym}: {type(e).__name__}: {str(e)[:50]}"); continue
        if not quarters:
            continue
        n_files += 1
        for q in quarters:
            rows_out.append((sym, q["period"], q["fiscal_label"]) + tuple(q.get(c) for c in QCOLS))
            n_q += 1

    cols = "symbol, period, fiscal_label, " + ", ".join(QCOLS)
    updates = ", ".join(f"{c}=EXCLUDED.{c}" for c in (["fiscal_label"] + QCOLS))
    execute_values(cur, f"""
        INSERT INTO quarterly_financials ({cols}) VALUES %s
        ON CONFLICT (symbol, period) DO UPDATE SET {updates}, updated_at=NOW()
    """, rows_out, page_size=500)
    conn.commit(); conn.close()
    print(f"quarterly_financials: wrote {n_q:,} quarter-rows across {n_files:,} files.")


if __name__ == "__main__":
    main()
