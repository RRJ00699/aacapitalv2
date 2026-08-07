#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""nse_preopen_capture.py — NSE IPO Special Pre-Open scraper (listing-day only).

ENDPOINT (repo-precedented, same family as fetch_nse_ipos FEEDS; wayback
harvester targets it; OWNER MUST browser-verify once before enabling cron —
ritual in docs/runbooks/NSE_PREOPEN_CAPTURE.md):
  page: https://www.nseindia.com/market-data/pre-open-market-cm-and-emerge-market
  api : https://www.nseindia.com/api/market-data-pre-open?key=IPO

Behavior:
- Self-exits in <1s when no mainboard IPO lists today (launcher pattern).
- Session: curl_cffi chrome124 impersonation + homepage/page priming +
  matching Referer (fetch_nse_ipos' proven Akamai-passing recipe). Re-primes
  on 401/403/parse failure; conservative 20s±3 jitter polling.
- Window: 08:59:30–10:00:30 IST only, then exits.
- First-contact SHAPE VALIDATION: required keys must exist or we abort to the
  failure sink (protects against NSE schema changes).
- Storage: raw JSON -> nse_preopen_raw (only when state changed); normalized ->
  ipo_preopen_book (source='nse-live') with state_hash UNIQUE for idempotent
  ON CONFLICT DO NOTHING dedupe. Fields normalized: IEP, IEQ, total buy/sell,
  best bid/ask (+qty from book levels), cancelled qty WHEN EXPOSED.
- Alerts: 3 consecutive failures or all-symbols-stale -> pipeline_failures
  (Pipeline health/phone) + best-effort ntfy push when NTFY_TOPIC set.
Run (LOCAL TEST MODE — no DB, Excel + raw JSONL; Rakesh 2026-07-17: prove it
locally first, graduate to DB or drop it later; data is for backtests):
  pip install curl_cffi openpyxl
  python _scripts/nse_preopen_capture.py --xlsx preopen.xlsx --once   # any time: shape check
  python _scripts/nse_preopen_capture.py --xlsx preopen.xlsx         # listing morning window
  (add --symbols ALPINE,SBIFM to filter; default captures ALL rows in the IPO set)
DB mode (dormant until approved): omit --xlsx on the VM.
"""
import argparse, datetime as dt, hashlib, json, os, random, sys, time
import psycopg2

NSE_BASE = "https://www.nseindia.com"
PAGE = f"{NSE_BASE}/market-data/pre-open-market-cm-and-emerge-market"
API = f"{NSE_BASE}/api/market-data-pre-open?key=IPO"
HEADERS = {"Accept": "application/json, text/plain, */*",
           "Accept-Language": "en-US,en;q=0.9", "Referer": PAGE}
POLL_S, JITTER_S = 20, 3

def ist_now():
    return dt.datetime.utcnow() + dt.timedelta(hours=5, minutes=30)

def in_window(t):
    a = t.replace(hour=8, minute=59, second=30, microsecond=0)
    b = t.replace(hour=10, minute=0, second=30, microsecond=0)
    return a <= t <= b

def db():
    return psycopg2.connect(os.environ["DATABASE_URL"])

def listings_today(cur):
    cur.execute("""SELECT UPPER(COALESCE(NULLIF(nse_symbol,''), symbol)) FROM ipo_intelligence
                   WHERE listing_date = (now() AT TIME ZONE 'Asia/Kolkata')::date
                     AND COALESCE(NULLIF(nse_symbol,''), symbol) IS NOT NULL""")
    return {r[0] for r in cur.fetchall()}

def fail(cur, note, alert=False):
    try:
        cur.execute("INSERT INTO pipeline_failures (step, script, stderr_tail) VALUES (%s,%s,%s)",
                    ("nse pre-open capture", "nse_preopen_capture.py", note[:490]))
        cur.connection.commit()
    except Exception:
        pass
    if alert and os.getenv("NTFY_TOPIC"):
        try:
            import urllib.request
            urllib.request.urlopen(urllib.request.Request(
                f"https://ntfy.sh/{os.environ['NTFY_TOPIC']}",
                data=("NSE pre-open capture: " + note[:180]).encode(),
                headers={"Title": "AACapital alert"}), timeout=10)
        except Exception:
            pass

def fail_safe(cur, note, alert=False):
    if cur is None:
        print(f"ALERT (local mode): {note}")
    else:
        fail(cur, note, alert)

def prime(cffi):
    s = cffi.Session(impersonate="chrome124")
    s.get(NSE_BASE, headers=HEADERS, timeout=12); time.sleep(0.8)
    s.get(PAGE, headers=HEADERS, timeout=12); time.sleep(0.8)
    return s

def validate_shape(payload):
    if not isinstance(payload, dict) or "data" not in payload:
        return "top-level 'data' missing"
    for row in payload.get("data") or []:
        md = row.get("metadata") or {}
        if "symbol" not in md:
            return "row metadata.symbol missing"
        break
    return None

def parse_rows(payload, want):
    out = []
    for row in payload.get("data") or []:
        md = row.get("metadata") or {}
        sym = str(md.get("symbol") or "").upper()
        if want and sym not in want:
            continue
        det = (row.get("detail") or {}).get("preOpenMarket") or {}
        levels = det.get("preopen") or []
        bids = [(l.get("price"), l.get("buyQty")) for l in levels if (l.get("buyQty") or 0) > 0 and l.get("price")]
        asks = [(l.get("price"), l.get("sellQty")) for l in levels if (l.get("sellQty") or 0) > 0 and l.get("price")]
        bb = max(bids, key=lambda x: x[0]) if bids else (None, None)
        ba = min(asks, key=lambda x: x[0]) if asks else (None, None)
        iep = md.get("lastPrice") or det.get("IEP")
        ieq = md.get("finalQuantity") or det.get("totalTradedQuantity")
        bq, sq = md.get("totalBuyQuantity") or det.get("totalBuyQuantity"), \
                 md.get("totalSellQuantity") or det.get("totalSellQuantity")
        lean = round((float(bq) - float(sq)) / float(sq) * 100, 1) if (bq and sq and float(sq) > 0) else None
        rec = {"symbol": sym, "iep": iep, "ieq": ieq, "buy_qty": bq, "sell_qty": sq,
               "lean_pct": lean, "best_bid": bb[0], "best_bid_qty": bb[1],
               "best_ask": ba[0], "best_ask_qty": ba[1],
               "cancelled_qty": md.get("cancelledQty") or det.get("cancelledQty"),
               "raw_row": row}
        rec["state_hash"] = hashlib.md5(
            f"{sym}|{iep}|{ieq}|{bq}|{sq}|{bb}|{ba}".encode()).hexdigest()
        out.append(rec)
    return out

def store(cur, rec):
    cur.execute("""INSERT INTO ipo_preopen_book
        (symbol, discovery_price, buy_qty, sell_qty, lean_pct, source,
         ieq, best_bid, best_bid_qty, best_ask, best_ask_qty, cancelled_qty, state_hash)
        VALUES (%s,%s,%s,%s,%s,'nse-live',%s,%s,%s,%s,%s,%s,%s)
        ON CONFLICT (state_hash) DO NOTHING""",
        (rec["symbol"], rec["iep"], rec["buy_qty"], rec["sell_qty"], rec["lean_pct"],
         rec["ieq"], rec["best_bid"], rec["best_bid_qty"], rec["best_ask"],
         rec["best_ask_qty"], rec["cancelled_qty"], rec["state_hash"]))
    if cur.rowcount:   # state changed -> also archive raw
        cur.execute("INSERT INTO nse_preopen_raw (symbol, payload) VALUES (%s,%s)",
                    (rec["symbol"], json.dumps(rec["raw_row"])))
    return cur.rowcount

def xlsx_store(path, rec, t):
    """Excel row + raw JSONL sidecar. Append-safe; header on create."""
    from openpyxl import Workbook, load_workbook
    cols = ["captured_ist", "symbol", "iep", "ieq", "buy_qty", "sell_qty", "lean_pct",
            "best_bid", "best_bid_qty", "best_ask", "best_ask_qty", "cancelled_qty", "state_hash"]
    if os.path.exists(path):
        wb = load_workbook(path); ws = wb.active
    else:
        wb = Workbook(); ws = wb.active; ws.title = "preopen"; ws.append(cols)
    ws.append([t.strftime("%Y-%m-%d %H:%M:%S")] + [rec[c] for c in cols[1:]])
    wb.save(path)
    with open(os.path.splitext(path)[0] + "_raw.jsonl", "a", encoding="utf-8") as f:
        f.write(json.dumps({"captured_ist": t.isoformat(), "symbol": rec["symbol"],
                            "raw": rec["raw_row"]}) + "\n")

def main():
    ap = argparse.ArgumentParser(); ap.add_argument("--once", action="store_true")
    ap.add_argument("--xlsx", help="LOCAL TEST MODE: write to this Excel file, no DB")
    ap.add_argument("--symbols", help="comma-separated filter (xlsx mode); default = all IPO-set rows")
    a = ap.parse_args()
    if a.xlsx:
        conn = cur = None
        want = {x.strip().upper() for x in a.symbols.split(",")} if a.symbols else None
        print(f"LOCAL XLSX MODE -> {a.xlsx} · symbols: {sorted(want) if want else 'ALL in IPO set'}")
    else:
        conn = db(); cur = conn.cursor()
        want = listings_today(cur)
        if not want:
            print("no mainboard IPO lists today — exit."); return
        print(f"listing-day capture for: {sorted(want)}")
    try:
        from curl_cffi import requests as cffi
    except ImportError:
        fail_safe(cur, "curl_cffi missing — pip install curl_cffi", alert=True); sys.exit(1)
    s = prime(cffi)
    consec, last_hashes, stale, seen = 0, {}, 0, set()
    while True:
        t = ist_now()
        if not a.once and not in_window(t):
            if t.hour >= 10: print("window closed — exit."); break
            time.sleep(15); continue
        try:
            r = s.get(API, headers=HEADERS, timeout=15)
            if r.status_code in (401, 403):
                raise RuntimeError(f"HTTP {r.status_code} (session/block)")
            payload = r.json()
            shape_err = validate_shape(payload)
            if shape_err:
                fail_safe(cur, f"SHAPE CHANGED: {shape_err}", alert=True); sys.exit(1)
            recs = parse_rows(payload, want)
            if a.xlsx:
                wrote = 0
                for rec in recs:
                    if rec["state_hash"] not in seen:
                        xlsx_store(a.xlsx, rec, t); seen.add(rec["state_hash"]); wrote += 1
            else:
                wrote = sum(store(cur, rec) for rec in recs)
                conn.commit()
            consec = 0
            hashes = {r0["symbol"]: r0["state_hash"] for r0 in recs}
            stale = stale + 1 if recs and hashes == last_hashes else 0
            last_hashes = hashes
            if stale == 8:
                fail_safe(cur, "book stale 8 consecutive polls during discovery — check page vs API")
            print(f"{t:%H:%M:%S} IST · {len(recs)} syms · {wrote} new states")
            if not recs:
                fail_safe(cur, "empty IPO pre-open set during window (symbols expected)")
        except SystemExit:
            raise
        except Exception as e:
            consec += 1
            print(f"poll error ({consec}): {str(e)[:100]}")
            if consec == 3:
                fail_safe(cur, f"3 consecutive failures: {str(e)[:200]}", alert=True)
            time.sleep(45)
            try: s = prime(cffi)
            except Exception: pass
        if a.once:
            break
        time.sleep(POLL_S + random.uniform(-JITTER_S, JITTER_S))
    if conn: conn.close()

if __name__ == "__main__":
    main()
