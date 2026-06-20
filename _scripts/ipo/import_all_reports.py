"""
_scripts/ipo/import_all_reports.py
=====================================
Imports ALL Chittorgarh Pro Excel reports into ipo_intelligence.

Files processed:
  Mainboard_IPOs_IPO_Subscription_Listing_Details_YYYY.xlsx  → core data
  Mainboard_IPOs_Listings_YYYY.xlsx                          → NSE/BSE symbols, ISIN
  Mainboard_IPOs_Anchor_Investor_Lock_in_Period_YYYY.xlsx    → T+30/T+90 dates
  Mainboard_IPOs_Key_Performance_Indicator_KPI_YYYY.xlsx     → ROE/ROCE/PE/EPS
  Lead_Managers_By_No_and_Performance_YYYY.xlsx              → BRLM track records

Usage:
  python _scripts/ipo/import_all_reports.py --dir data/ipo_reports
  python _scripts/ipo/import_all_reports.py --dir data/ipo_reports --dry-run
"""

import os, sys, re, math, json, logging, argparse, datetime
import openpyxl, psycopg2, psycopg2.extras
from collections import defaultdict

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
log = logging.getLogger(__name__)

DATABASE_URL = os.environ.get("DATABASE_URL") or os.environ.get("NEON_DATABASE_URL")

TIER1_ANCHORS = {
    "lic","life insurance","sbi mutual","sbi mf","icici prudential","icici pru",
    "nippon","hdfc mutual","hdfc mf","kotak mutual","kotak mf","adia","abu dhabi",
    "gic","singapore","norway","temasek","axis mutual","axis mf","dsp","mirae",
    "franklin","motilal","canara robeco","tata mutual","tata mf","aditya birla",
    "pgim","sundaram","uti mutual","uti mf",
}

def get_db():
    return psycopg2.connect(DATABASE_URL, connect_timeout=15)

def n(v, d=None):
    try:
        if v is None: return d
        if isinstance(v,(int,float)):
            return None if math.isnan(float(v)) else float(v)
        s = str(v).split('(')[0].strip().replace(',','').replace('₹','').replace('Rs.','')
        return float(s) if s else d
    except: return d

def parse_date(v):
    if v is None: return None
    if isinstance(v,(datetime.date,datetime.datetime)):
        return v.date() if hasattr(v,'date') else v
    s = str(v).strip()
    if s in ('','NaT','nan'): return None
    # Handle "2024-12-31 00:00:00" format from openpyxl
    s = s[:10]
    for fmt in ('%Y-%m-%d','%d-%b-%Y','%d/%m/%Y','%d %b %Y'):
        try: return datetime.datetime.strptime(s, fmt).date()
        except: continue
    return None

def read_file(path):
    """Returns (headers, data_rows) for a Chittorgarh Excel file."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    for i, row in enumerate(rows):
        first = str(row[0] or '').strip()
        if first in ('Company','Lead Manager','Anchor Investor','Rank','Registrar'):
            return [str(c or '').strip() for c in row], rows[i+1:]
    return None, []

def compute_operator_risk(ipo: dict) -> float:
    score = 0.0
    size = n(ipo.get('issue_size_cr')) or 0
    if 0 < size < 150:   score += 40
    elif size < 300:     score += 15
    qib = n(ipo.get('qib_subscription_x')) or 0
    ret = n(ipo.get('rii_subscription_x')) or 0
    if qib < 5 and ret > 50: score += 15
    return min(score, 100)

def compute_play(ipo: dict) -> dict:
    size   = n(ipo.get('issue_size_cr')) or 0
    qib    = n(ipo.get('qib_subscription_x')) or 0
    ret_op = n(ipo.get('return_listing_open')) or 0
    anc_t1 = int(n(ipo.get('anchor_tier1_count')) or 0)
    op_r   = n(ipo.get('operator_risk_score')) or 0

    if size > 0 and size < 150:
        return {"play":"AVOID","confidence":90,
                "reasons":["Issue size < ₹150 Cr — 5% band, operator trap"],
                "stop_loss_pct":0,"target_pct":0,"hold_window":"—"}

    if op_r > 70:
        return {"play":"AVOID","confidence":80,
                "reasons":[f"High operator risk ({op_r:.0f})"],
                "stop_loss_pct":0,"target_pct":0,"hold_window":"—"}

    # We have actual returns — classify historical play
    if ret_op != 0:
        if ret_op < -5 and anc_t1 >= 8:
            return {"play":"BUY_PANIC_DIP","confidence":74,
                    "reasons":[f"Listed {ret_op:.1f}% — panic sell, {anc_t1} tier-1 anchors = support floor"],
                    "stop_loss_pct":6,"target_pct":20,"hold_window":"Day 1-3"}
        if ret_op > 25 and qib < 20:
            return {"play":"AVOID","confidence":75,
                    "reasons":[f"Listed +{ret_op:.0f}% with weak QIB ({qib:.0f}x) — euphoria trap"],
                    "stop_loss_pct":0,"target_pct":0,"hold_window":"—"}
        if ret_op > 0 and qib >= 30:
            return {"play":"BUY_AT_OPEN","confidence":78,
                    "reasons":[f"QIB {qib:.0f}x + listed +{ret_op:.0f}% — strong institutional demand"],
                    "stop_loss_pct":4,"target_pct":ret_op+8,"hold_window":"30min → EOD"}
        if ret_op > 0:
            return {"play":"WAIT_FOR_VWAP","confidence":65,
                    "reasons":[f"Listed +{ret_op:.0f}% — confirm VWAP before entry"],
                    "stop_loss_pct":5,"target_pct":ret_op+5,"hold_window":"10:30AM → EOD"}

    # Pre-listing
    if qib >= 50 and anc_t1 >= 15:
        return {"play":"BUY_AT_OPEN","confidence":82,
                "reasons":[f"QIB {qib:.0f}x, {anc_t1} tier-1 anchors — strong conviction"],
                "stop_loss_pct":4,"target_pct":22,"hold_window":"30min → EOD"}
    if qib >= 20 and anc_t1 >= 8:
        return {"play":"WAIT_FOR_VWAP","confidence":67,
                "reasons":[f"QIB {qib:.0f}x — wait for VWAP confirmation"],
                "stop_loss_pct":5,"target_pct":15,"hold_window":"10:30AM → EOD"}

    return {"play":"AVOID","confidence":58,
            "reasons":["Insufficient data or conviction for entry"],
            "stop_loss_pct":0,"target_pct":0,"hold_window":"—"}

def ensure_columns(conn):
    cols = [
        ("nse_symbol","TEXT"),("bse_code","TEXT"),("isin","TEXT"),
        ("listing_exchange","TEXT"),("listing_close_price","NUMERIC"),
        ("return_listing_open","NUMERIC"),("return_current","NUMERIC"),
        ("anchor_investment_cr","NUMERIC"),("anchor_alloc_pct","NUMERIC"),
        ("anchor_lock30_date","DATE"),("anchor_lock90_date","DATE"),
        ("anchor_tier1_count","INTEGER"),
        ("roe","NUMERIC"),("roce","NUMERIC"),("pat_margin","NUMERIC"),
        ("ebitda_pct","NUMERIC"),("ipo_pb","NUMERIC"),
        ("eps_pre","NUMERIC"),("eps_post","NUMERIC"),
        ("ipo_pe_pre","NUMERIC"),("ipo_pe_post","NUMERIC"),
        ("operator_risk_score","NUMERIC"),
        ("play_recommendation","TEXT"),("play_confidence","NUMERIC"),
        ("play_reasons","JSONB"),("play_stop_loss_pct","NUMERIC"),
        ("play_target_pct","NUMERIC"),("play_hold_window","TEXT"),
        ("play_updated_at","TIMESTAMPTZ"),
        ("chittorgarh_imported","BOOLEAN DEFAULT FALSE"),
        ("revenue_cr","NUMERIC"),("pat_cr","NUMERIC"),
        ("ebitda_cr","NUMERIC"),("net_worth_cr","NUMERIC"),
        ("total_debt_cr","NUMERIC"),("anchor_qib_alloc_cr","NUMERIC"),
    ]
    cur = conn.cursor()
    for col, typ in cols:
        try:
            cur.execute(f"ALTER TABLE ipo_intelligence ADD COLUMN IF NOT EXISTS {col} {typ}")
            conn.commit()
        except: conn.rollback()
    cur.close()

def upsert_ipo(conn, data: dict):
    cur = conn.cursor()
    company = str(data.get('company_name','')).strip()
    if not company: return False

    data['chittorgarh_imported'] = True
    data['play_updated_at'] = datetime.datetime.now(datetime.timezone.utc)

    cur.execute("SELECT id FROM ipo_intelligence WHERE company_name = %s LIMIT 1", (company,))
    existing = cur.fetchone()

    cols = [c for c in data.keys() if c != 'company_name']
    vals = [data[c] for c in cols]

    if existing:
        set_clause = ', '.join([f"{c} = %s" for c in cols])
        cur.execute(f"UPDATE ipo_intelligence SET {set_clause} WHERE company_name = %s",
                    vals + [company])
    else:
        all_cols = ['company_name'] + cols
        all_vals = [company] + vals
        ph = ', '.join(['%s'] * len(all_vals))
        cur.execute(f"INSERT INTO ipo_intelligence ({', '.join(all_cols)}) VALUES ({ph})", all_vals)

    conn.commit()
    cur.close()
    return True

def upsert_brlm(conn, brlm_scores: dict):
    cur = conn.cursor()
    cur.execute("""
        CREATE TABLE IF NOT EXISTS brlm_scores (
            brlm_name TEXT PRIMARY KEY, score NUMERIC,
            ipo_count INTEGER, avg_listing NUMERIC,
            avg_day30 NUMERIC, pct_negative NUMERIC,
            updated_at TIMESTAMPTZ DEFAULT NOW()
        )
    """)
    conn.commit()
    for brlm, s in brlm_scores.items():
        cur.execute("""
            INSERT INTO brlm_scores (brlm_name, score, ipo_count, avg_listing, pct_negative, updated_at)
            VALUES (%s,%s,%s,%s,%s,NOW())
            ON CONFLICT (brlm_name) DO UPDATE SET
                score=EXCLUDED.score, ipo_count=EXCLUDED.ipo_count,
                avg_listing=EXCLUDED.avg_listing, pct_negative=EXCLUDED.pct_negative,
                updated_at=NOW()
        """, (brlm, s['score'], s['ipo_count'], s['avg_listing'], s['pct_negative']))
    conn.commit()
    cur.close()

def main():
    p = argparse.ArgumentParser()
    p.add_argument("--dir",     required=True, help="Directory with Chittorgarh Excel files")
    p.add_argument("--dry-run", action="store_true")
    args = p.parse_args()

    if not os.path.isdir(args.dir):
        log.error(f"Directory not found: {args.dir}"); sys.exit(1)

    if not DATABASE_URL and not args.dry_run:
        log.error("DATABASE_URL not set"); sys.exit(1)

    log.info(f"Loading all reports from {args.dir}")
    log.info("=" * 60)

    ipos = {}

    # ── 1. Subscription + Listing (core) ─────────────────────────────────────
    for year in range(2021, 2027):
        f = f"{args.dir}/Mainboard_IPOs_IPO_Subscription_Listing_Details_{year}.xlsx"
        if not os.path.exists(f): continue
        headers, rows = read_file(f)
        if not headers: continue
        count = 0
        for row in rows:
            if not row[0] or str(row[0]).strip() in ('','Company'): continue
            company = str(row[0]).strip()
            vals = [str(c or '').strip() for c in row]
            ipos[company] = {
                'company_name':         company,
                'is_sme':               False,
                'open_date':            parse_date(vals[1]) if len(vals)>1 else None,
                'issue_price':          n(vals[2]) if len(vals)>2 else None,
                'issue_size_cr':        n(vals[3]) if len(vals)>3 else None,
                'qib_subscription_x':   n(vals[4]) if len(vals)>4 else None,
                'nii_subscription_x':   n(vals[5]) if len(vals)>5 else None,
                'rii_subscription_x':   n(vals[6]) if len(vals)>6 else None,
                'total_subscription_x': n(vals[10]) if len(vals)>10 else None,
                'listing_date':         parse_date(vals[11]) if len(vals)>11 else None,
                'listing_open':         n(vals[12]) if len(vals)>12 else None,
                'listing_close_price':  n(vals[13]) if len(vals)>13 else None,
                'return_listing_open':  n(vals[14]) if len(vals)>14 else None,
            }
            count += 1
        log.info(f"  Subscription {year}: {count} IPOs")

    log.info(f"  Total core: {len(ipos)} IPOs")

    # ── 2. Listings (symbols, ISIN, current price) ────────────────────────────
    enrich_count = 0
    for year in range(2021, 2027):
        f = f"{args.dir}/Mainboard_IPOs_Listings_{year}.xlsx"
        if not os.path.exists(f): continue
        headers, rows = read_file(f)
        if not headers: continue
        for row in rows:
            company = str(row[0] or '').strip()
            if not company or company == 'Company': continue
            vals = [str(c or '').strip() for c in row]
            if company in ipos:
                ipos[company].update({
                    'listing_exchange': vals[3] if len(vals)>3 else None,
                    'isin':             vals[4] if len(vals)>4 else None,
                    'bse_code':         vals[5] if len(vals)>5 else None,
                    'nse_symbol':       vals[6] if len(vals)>6 else None,
                    'return_current':   n(vals[12]) if len(vals)>12 else None,
                })
                enrich_count += 1
    log.info(f"  Listings enriched: {enrich_count}")

    # ── 3. Anchor lock-in dates ───────────────────────────────────────────────
    anchor_count = 0
    for year in range(2021, 2027):
        f = f"{args.dir}/Mainboard_IPOs_Anchor_Investor_Lock_in_Period_{year}.xlsx"
        if not os.path.exists(f): continue
        headers, rows = read_file(f)
        if not headers: continue
        for row in rows:
            company = str(row[0] or '').strip()
            if not company or company == 'Company': continue
            vals = [str(c or '').strip() for c in row]
            if company in ipos:
                ipos[company].update({
                    'anchor_investment_cr': n(vals[3]) if len(vals)>3 else None,
                    'anchor_alloc_pct':     n(vals[4]) if len(vals)>4 else None,
                    'anchor_lock30_date':   parse_date(vals[5]) if len(vals)>5 else None,
                    'anchor_lock90_date':   parse_date(vals[6]) if len(vals)>6 else None,
                })
                anchor_count += 1
    log.info(f"  Anchor dates: {anchor_count}")

    # ── 4. KPIs ───────────────────────────────────────────────────────────────
    kpi_count = 0
    for year in [2021,2022,2023,2024,2025,2026]:
        f = f"{args.dir}/Mainboard_IPOs_Key_Performance_Indicator_KPI_{year}.xlsx"
        if not os.path.exists(f): continue
        headers, rows = read_file(f)
        if not headers: continue
        for row in rows:
            company = str(row[0] or '').strip()
            if not company or company == 'Company': continue
            vals = [str(c or '').strip() for c in row]
            if company in ipos:
                ipos[company].update({
                    'roe':        n(vals[5]) if len(vals)>5 else None,
                    'roce':       n(vals[6]) if len(vals)>6 else None,
                    'pat_margin': n(vals[8]) if len(vals)>8 else None,
                    'ebitda_pct': n(vals[9]) if len(vals)>9 else None,
                    'ipo_pb':     n(vals[10]) if len(vals)>10 else None,
                    'eps_pre':    n(vals[11]) if len(vals)>11 else None,
                    'eps_post':   n(vals[12]) if len(vals)>12 else None,
                    'ipo_pe_pre': n(vals[13]) if len(vals)>13 else None,
                    'ipo_pe_post':n(vals[14]) if len(vals)>14 else None,
                })
                kpi_count += 1
    log.info(f"  KPIs: {kpi_count}")

    # ── 5. Key Financial Details ─────────────────────────────────
    fin_count = 0
    for year in [2021,2022,2023,2024,2025,2026]:
        f = f"{args.dir}/Mainboard_IPOs_Key_Financial_Details_{year}.xlsx"
        if not os.path.exists(f): continue
        headers, rows = read_file(f)
        if not headers: continue
        for row in rows:
            company = str(row[0] or '').strip()
            if not company or company == 'Company': continue
            vals = [str(c or '').strip() for c in row]
            if company in ipos:
                ipos[company].update({
                    'revenue_cr':    n(vals[5]) if len(vals)>5 else None,
                    'pat_cr':        n(vals[6]) if len(vals)>6 else None,
                    'ebitda_cr':     n(vals[7]) if len(vals)>7 else None,
                    'net_worth_cr':  n(vals[8]) if len(vals)>8 else None,
                    'total_debt_cr': n(vals[10]) if len(vals)>10 else None,
                })
                fin_count += 1
    log.info(f"  Financials: {fin_count}")

    # ── 6. BRLM track records ─────────────────────────────────────────────────
    brlm_raw = defaultdict(lambda: {'issues':0,'positive':0,'negative':0,'gains':[]})
    for year in range(2021, 2027):
        f = f"{args.dir}/Lead_Managers_By_No_and_Performance_of_Mainboard_IPOs_managed_{year}.xlsx"
        if not os.path.exists(f): continue
        headers, rows = read_file(f)
        if not headers: continue
        for row in rows:
            brlm = str(row[0] or '').strip()
            if not brlm or brlm in ('Lead Manager',''): continue
            vals = [str(c or '').strip() for c in row]
            brlm_raw[brlm]['issues']   += int(n(vals[1]) or 0)
            brlm_raw[brlm]['positive'] += int(n(vals[3]) or 0)
            brlm_raw[brlm]['negative'] += int(n(vals[5]) or 0)
            g = n(vals[7])
            if g: brlm_raw[brlm]['gains'].append(g)

    brlm_scores = {}
    for brlm, d in brlm_raw.items():
        total = d['issues'] or 1
        avg_g = sum(d['gains'])/len(d['gains']) if d['gains'] else 0
        pct_neg = d['negative']/total*100
        # Score: avg listing gain (weighted) + pct positive
        score = min(100, max(0,
            avg_g * 2.5 +           # 40% listing gain = 100
            (100 - pct_neg) * 0.6   # 0% negative = 60 pts
        ))
        brlm_scores[brlm] = {
            'score':      round(score, 1),
            'ipo_count':  total,
            'avg_listing':round(avg_g, 2),
            'pct_negative':round(pct_neg, 1),
        }
    log.info(f"  BRLM scores: {len(brlm_scores)}")

    # ── Compute derived fields for each IPO ───────────────────────────────────
    for company, ipo in ipos.items():
        ipo['operator_risk_score'] = compute_operator_risk(ipo)
        play = compute_play(ipo)
        ipo.update({
            'play_recommendation': play['play'],
            'play_confidence':     play['confidence'],
            'play_reasons':        json.dumps(play['reasons']),
            'play_stop_loss_pct':  play['stop_loss_pct'],
            'play_target_pct':     play['target_pct'],
            'play_hold_window':    play['hold_window'],
        })

    if args.dry_run:
        log.info("\nDRY RUN — sample:")
        for company, ipo in list(ipos.items())[:5]:
            log.info(f"  {company[:45]:45s} sub={ipo.get('total_subscription_x')}x "
                     f"listing={ipo.get('return_listing_open')}% "
                     f"play={ipo.get('play_recommendation')}")
        log.info(f"\nTop 5 BRLMs:")
        for brlm, s in sorted(brlm_scores.items(), key=lambda x:-x[1]['score'])[:5]:
            log.info(f"  {s['score']:5.1f}  {brlm[:45]} ({s['ipo_count']} IPOs, avg {s['avg_listing']:.1f}%)")
        return

    # ── Push to Neon ──────────────────────────────────────────────────────────
    conn = get_db()
    log.info("\nConnected to Neon DB")
    ensure_columns(conn)

    ok = 0; skipped = 0
    for company, ipo in ipos.items():
        try:
            if upsert_ipo(conn, {**ipo}):
                ok += 1
                if ok % 50 == 0:
                    log.info(f"  {ok}/{len(ipos)} imported…")
        except Exception as e:
            log.warning(f"  {company}: {e}")
            skipped += 1

    log.info(f"\n✅ IPOs: {ok} imported, {skipped} skipped")

    # Push BRLM scores
    upsert_brlm(conn, brlm_scores)
    log.info(f"✅ BRLM: {len(brlm_scores)} scores saved")

    conn.close()
    log.info("=" * 60)
    log.info(f"Complete. {ok} IPOs + {len(brlm_scores)} BRLM scores in Neon")

if __name__ == "__main__":
    main()
