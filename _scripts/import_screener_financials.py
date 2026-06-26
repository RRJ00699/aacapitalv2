#!/usr/bin/env python3
"""
import_screener_financials.py — load the FULL 10-year financials from Screener Excel
exports (the *_10yr.xlsx files in data/fundamental_raw/) into annual_financials.

Your existing screener-pipeline.ts downloads these and stuffs only revenue/PAT into
quarterly_results — the Balance Sheet and Cash Flow it parses are discarded. This loads
every section (full P&L line items, complete Balance Sheet, Cash Flow, share count) so the
Financial-DNA / balance-sheet engine has something to read, and so EPS can be derived
(net_profit / no_of_shares).

Source layout (Data Sheet): sections PROFIT & LOSS / BALANCE SHEET / CASH FLOW: each carry
10 fiscal years aligned on a 'Report Date' row; values in columns B.. . Annual only here —
quarterly stays in quarterly_results.

Usage:
  python _scripts/import_screener_financials.py --dir "C:\\aacapital-v2\\data\\fundamental_raw"
  python _scripts/import_screener_financials.py --dir ./data/fundamental_raw --diag   # parse 1 file, print, no DB
Env: DATABASE_URL (or NEON_DATABASE_URL)
"""
import os, sys, glob, argparse, datetime
import openpyxl

URL = os.environ.get("DATABASE_URL") or os.environ.get("NEON_DATABASE_URL")

# section header (col A) -> {screener label: db_column}; order matters for duplicate 'Total'
PL = {
    "Sales": "sales", "Raw Material Cost": "raw_material_cost", "Change in Inventory": "change_in_inventory",
    "Power and Fuel": "power_fuel", "Other Mfr. Exp": "other_mfr_exp", "Employee Cost": "employee_cost",
    "Selling and admin": "selling_admin", "Other Expenses": "other_expenses", "Other Income": "other_income",
    "Depreciation": "depreciation", "Interest": "interest", "Profit before tax": "pbt",
    "Tax": "tax", "Net profit": "net_profit", "Dividend Amount": "dividend_amount",
}
BS = {
    "Equity Share Capital": "equity_capital", "Reserves": "reserves", "Borrowings": "borrowings",
    "Other Liabilities": "other_liabilities", "Net Block": "net_block",
    "Capital Work in Progress": "cwip", "Investments": "investments", "Other Assets": "other_assets",
    "Receivables": "receivables", "Inventory": "inventory", "Cash & Bank": "cash_bank",
    "No. of Equity Shares": "no_of_shares", "Face value": "face_value",
    # the two bare 'Total' rows handled positionally below
}
CF = {
    "Cash from Operating Activity": "cfo", "Cash from Investing Activity": "cfi",
    "Cash from Financing Activity": "cff", "Net Cash Flow": "net_cash_flow",
}

ALL_COLS = (["report_date"] + list(PL.values())
            + ["total_liabilities", "total_assets"] + list(BS.values())
            + list(CF.values()))


def num(v):
    if isinstance(v, (int, float)):
        return float(v)
    return None


def parse_file(path):
    """Return {fiscal_year(int): {db_column: value}} merged across P&L/BS/CF."""
    wb = openpyxl.load_workbook(path, data_only=True)
    if "Data Sheet" not in wb.sheetnames:
        return None, None
    ws = wb["Data Sheet"]
    rows = [[ws.cell(r, c).value for c in range(1, ws.max_column + 1)] for r in range(1, ws.max_row + 1)]

    company = None
    for rr in rows:
        if rr and rr[0] == "COMPANY NAME":
            company = rr[1]
            break

    def section_bounds(header):
        for i, rr in enumerate(rows):
            if rr and rr[0] == header:
                return i
        return None

    # Sections run P&L -> Quarters -> Balance Sheet -> Cash Flow. Quarters repeats the same
    # labels (Sales, Net profit, ...), so each section MUST stop at the next header — otherwise
    # P&L would sweep through Quarters and the quarterly rows overwrite the annual ones.
    ALL_HEADERS = ["META", "PROFIT & LOSS", "Quarters", "BALANCE SHEET", "CASH FLOW:", "PRICE:", "DERIVED:"]
    hdr_pos = sorted((p, h) for h in ALL_HEADERS for p in [section_bounds(h)] if p is not None)

    years = {}   # fiscal_year -> dict

    def read_section(header, mapping, _next=None):
        start = section_bounds(header)
        if start is None:
            return
        end = len(rows)
        for p, h in hdr_pos:                 # stop at the next section header, whatever it is
            if p > start:
                end = p
                break
        # find Report Date row in this section → fiscal-year columns
        dates = None
        for i in range(start, end):
            if rows[i] and rows[i][0] == "Report Date":
                dates = rows[i][1:]
                break
        if not dates:
            return
        fys = []
        for d in dates:
            if isinstance(d, datetime.datetime):
                fys.append(d.year)
            else:
                fys.append(None)
        total_seen = 0
        for i in range(start, end):
            label = rows[i][0] if rows[i] else None
            if label is None:
                continue
            col = mapping.get(label)
            if col is None and label == "Total" and header == "BALANCE SHEET":
                col = "total_liabilities" if total_seen == 0 else "total_assets"
                total_seen += 1
            if col is None:
                continue
            vals = rows[i][1:]
            for j, fy in enumerate(fys):
                if fy is None or j >= len(vals):
                    continue
                rec = years.setdefault(fy, {})
                rec[col] = num(vals[j])
                if col == "report_date" or "report_date" not in rec:
                    rec["report_date"] = dates[j] if isinstance(dates[j], datetime.datetime) else None

    read_section("PROFIT & LOSS", PL, "BALANCE SHEET")
    read_section("BALANCE SHEET", BS, "CASH FLOW:")
    read_section("CASH FLOW:", CF, None)
    return company, years


def symbol_from_path(path):
    base = os.path.basename(path)
    return base.replace("_10yr.xlsx", "").replace(".xlsx", "").upper()


def ensure_table(cur):
    cols_sql = ",\n            ".join(f"{c} NUMERIC(18,4)" for c in ALL_COLS if c != "report_date")
    cur.execute(f"""
        CREATE TABLE IF NOT EXISTS annual_financials (
            symbol        TEXT NOT NULL,
            fiscal_year   INT  NOT NULL,
            report_date   DATE,
            company_name  TEXT,
            {cols_sql},
            source        TEXT DEFAULT 'screener',
            updated_at    TIMESTAMPTZ DEFAULT NOW(),
            UNIQUE (symbol, fiscal_year)
        )
    """)
    for c in ALL_COLS:
        if c == "report_date":
            continue
        cur.execute(f"ALTER TABLE annual_financials ADD COLUMN IF NOT EXISTS {c} NUMERIC(18,4)")


SCREENER_BASE = "https://www.screener.in"


def download_one(symbol, out_dir, cookie, ua, timeout=30):
    """Download {symbol}_10yr.xlsx from Screener. Returns (status, reason).
    status: 'exists' | 'downloaded' | 'failed'. Mirrors screener-pipeline.ts."""
    import requests
    out = os.path.join(out_dir, f"{symbol}_10yr.xlsx")
    if os.path.exists(out) and os.path.getsize(out) > 5000:
        return "exists", ""
    headers = {"Cookie": cookie, "User-Agent": ua}
    reason = "no response"
    for url in (f"{SCREENER_BASE}/company/{symbol}/consolidated/export/",
                f"{SCREENER_BASE}/company/{symbol}/export/"):
        headers["Referer"] = url.rsplit("export/", 1)[0]
        try:
            r = requests.get(url, headers=headers, timeout=timeout)
        except Exception as e:
            reason = f"{type(e).__name__}: {str(e)[:40]}"
            continue
        if r.status_code != 200:
            snippet = r.text[:120].replace("\n", " ") if hasattr(r, "text") else ""
            reason = f"HTTP {r.status_code} | {snippet}"     # 404 bad slug · 403/302 auth · 'cloudflare'=challenge
            continue
        if r.content[:4].hex() == "504b0304":                # real .xlsx (zip) signature
            with open(out, "wb") as f:
                f.write(r.content)
            return "downloaded", ""
        reason = f"not xlsx | {r.text[:120].replace(chr(10),' ')}"
    return "failed", reason


def fetch_symbols(args, cookie):
    """Resolve symbols to download from --symbol / --symbols-file / company_master."""
    if args.symbol:
        return [args.symbol.upper()]
    if args.symbols_file:
        with open(args.symbols_file, encoding="utf-8") as f:
            return [ln.strip().upper() for ln in f if ln.strip()]
    if not URL:
        sys.exit("To --download the whole universe set DATABASE_URL (reads company_master), "
                 "or pass --symbols-file / --symbol.")
    import psycopg2
    conn = psycopg2.connect(URL); cur = conn.cursor()
    cur.execute("SELECT symbol FROM company_master WHERE symbol IS NOT NULL ORDER BY symbol")
    syms = [r[0].upper() for r in cur.fetchall()]
    cur.close(); conn.close()
    return syms


def download_phase(args, cookie, ua):
    import time
    os.makedirs(args.dir, exist_ok=True)
    syms = fetch_symbols(args, cookie)
    if args.limit:
        syms = syms[:args.limit]
    print(f"Download phase: {len(syms):,} symbols (resume skips existing).")
    counts = {"downloaded": 0, "exists": 0, "failed": 0}
    fails = []
    for i, sym in enumerate(syms, 1):
        status, reason = download_one(sym, args.dir, cookie, ua)
        counts[status] += 1
        if status == "downloaded":
            print(f"  [{i}/{len(syms)}] {sym}  downloaded")
        elif status == "failed":
            fails.append((sym, reason))
            print(f"  [{i}/{len(syms)}] {sym}  FAILED - {reason}")
        if status != "exists":
            time.sleep(args.sleep)          # throttle only on real network hits
    print(f"\nDownloaded {counts['downloaded']:,} | already had {counts['exists']:,} | failed {counts['failed']:,}")
    if fails:
        fpath = os.path.join(args.dir, "_download_failures.txt")
        with open(fpath, "w", encoding="utf-8") as f:
            f.write("\n".join(f"{s}\t{r}" for s, r in fails) + "\n")
        print(f"Failure reasons -> {fpath}")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--dir", required=True)
    ap.add_argument("--diag", action="store_true", help="parse one file, print, no DB writes")
    ap.add_argument("--download", action="store_true", help="fetch missing files from Screener first")
    ap.add_argument("--cookie", help="Screener cookie; or set SCREENER_COOKIE env")
    ap.add_argument("--symbols-file", dest="symbols_file", help="newline list of symbols to fetch")
    ap.add_argument("--symbol", help="fetch a single symbol")
    ap.add_argument("--limit", type=int, default=0, help="cap how many symbols to fetch")
    ap.add_argument("--sleep", type=float, default=1.5, help="seconds between downloads (rate-limit guard)")
    ap.add_argument("--user-agent", dest="user_agent",
                    default="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
                    help="MUST match the browser that produced cf_clearance")
    args = ap.parse_args()

    if args.download:
        cookie = os.environ.get("SCREENER_COOKIE") or args.cookie
        if not cookie:
            sys.exit("Set the cookie: $env:SCREENER_COOKIE='csrftoken=...; sessionid=...'  (or --cookie '...')")
        try:
            import requests  # noqa: F401
        except ImportError:
            sys.exit("pip install requests")
        download_phase(args, cookie, args.user_agent)

    files = sorted(glob.glob(os.path.join(args.dir, "*.xlsx")))
    if not files:
        sys.exit(f"No .xlsx files in {args.dir}")
    print(f"Found {len(files):,} files.")

    if args.diag:
        company, years = parse_file(files[0])
        print(f"\n{symbol_from_path(files[0])} ({company}) — {len(years)} fiscal years")
        for fy in sorted(years):
            r = years[fy]
            print(f"  FY{fy}: sales={r.get('sales')}, net_profit={r.get('net_profit')}, "
                  f"borrowings={r.get('borrowings')}, reserves={r.get('reserves')}, "
                  f"cfo={r.get('cfo')}, shares={r.get('no_of_shares')}, "
                  f"total_assets={r.get('total_assets')}")
        miss = [c for c in ALL_COLS if all(c not in years[fy] or years[fy][c] is None for fy in years)]
        print(f"  columns never populated: {miss or 'none'}")
        return

    if not URL:
        sys.exit("DATABASE_URL not set")
    import psycopg2
    from psycopg2.extras import execute_values
    conn = psycopg2.connect(URL); conn.autocommit = False
    cur = conn.cursor()
    ensure_table(cur); conn.commit()

    payload, n_years, n_files = [], 0, 0
    data_cols = [c for c in ALL_COLS if c != "report_date"]
    for path in files:
        sym = symbol_from_path(path)
        try:
            company, years = parse_file(path)
        except Exception as e:
            print(f"  ✗ {sym}: {e}")
            continue
        if not years:
            print(f"  ⚠ {sym}: no annual data parsed")
            continue
        n_files += 1
        for fy in sorted(years):
            r = years[fy]
            rd = r.get("report_date")
            row = [sym, fy, rd.date() if isinstance(rd, datetime.datetime) else None, company]
            row += [r.get(c) for c in data_cols]
            payload.append(tuple(row))
            n_years += 1

    insert_cols = ["symbol", "fiscal_year", "report_date", "company_name"] + data_cols
    update_set = ", ".join(f"{c}=EXCLUDED.{c}" for c in ["report_date", "company_name"] + data_cols)
    execute_values(cur, f"""
        INSERT INTO annual_financials ({", ".join(insert_cols)})
        VALUES %s
        ON CONFLICT (symbol, fiscal_year) DO UPDATE SET {update_set}, updated_at=NOW()
    """, payload, page_size=500)
    conn.commit()
    print(f"annual_financials: wrote {n_years:,} company-years across {n_files:,} files.")
    cur.close(); conn.close()


if __name__ == "__main__":
    main()
